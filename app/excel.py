from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .catalog import PERIOD_LABELS


HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)


def cell_import_value(cell):
    value = cell.value
    if isinstance(value, (int, float)) and "%" in str(cell.number_format):
        return value * 100
    return value


def create_report_template(report_key: str, definition: dict, period_type: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Итоги"

    sheet["A1"] = "Код отчета"
    sheet["B1"] = report_key
    sheet["A2"] = "Тип периода"
    sheet["B2"] = period_type
    sheet["A3"] = "Название"
    sheet["B3"] = definition["title"]
    sheet["A4"] = "Период"
    sheet["B4"] = PERIOD_LABELS[period_type]
    sheet["A5"] = "Заполняйте только значения и названия артикулов. После импорта в систему сохраняются только данные, а не сам Excel-файл."
    sheet["A5"].font = TITLE_FONT

    headers = ["Код", "Показатель", "Единица", "План", "Факт", "Комментарий"]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=7, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_index, metric in enumerate(definition["metrics"], start=8):
        sheet.cell(row=row_index, column=1, value=metric["code"])
        sheet.cell(row=row_index, column=2, value=metric["label"])
        sheet.cell(row=row_index, column=3, value=metric["unit"])
        if not metric["has_plan"]:
            sheet.cell(row=row_index, column=4, value="")
        sheet.cell(row=row_index, column=5, value="")
        sheet.cell(row=row_index, column=6, value="")

    sheet.column_dimensions["A"].hidden = True
    widths = [18, 42, 12, 18, 18, 34]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    article_sheet = workbook.create_sheet("Поартикульно")
    article_sheet["A1"] = "Артикул"
    article_sheet["B1"] = "Код показателя"
    article_sheet["C1"] = "Показатель"
    article_sheet["D1"] = "План"
    article_sheet["E1"] = "Факт"
    article_sheet["F1"] = "Комментарий"
    for column in range(1, 7):
        cell = article_sheet.cell(row=1, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_index, metric in enumerate(definition["metrics"], start=2):
        article_sheet.cell(row=row_index, column=2, value=metric["code"])
        article_sheet.cell(row=row_index, column=3, value=metric["label"])

    article_widths = [28, 18, 36, 18, 18, 34]
    for idx, width in enumerate(article_widths, start=1):
        article_sheet.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def parse_report_workbook(file_storage) -> dict:
    workbook = load_workbook(file_storage, data_only=True)
    sheet = workbook["Итоги"] if "Итоги" in workbook.sheetnames else workbook.active
    metrics = []
    row = 8
    while True:
        code = sheet.cell(row=row, column=1).value
        label = sheet.cell(row=row, column=2).value
        if not code and not label:
            break
        metrics.append(
            {
                "code": str(code).strip(),
                "plan": cell_import_value(sheet.cell(row=row, column=4)),
                "fact": cell_import_value(sheet.cell(row=row, column=5)),
                "comment": sheet.cell(row=row, column=6).value,
            }
        )
        row += 1

    article_entries = []
    if "Поартикульно" in workbook.sheetnames:
        article_sheet = workbook["Поартикульно"]
        current_article_name = None
        for row in range(2, article_sheet.max_row + 1):
            article_name_cell = article_sheet.cell(row=row, column=1).value
            if article_name_cell not in (None, ""):
                current_article_name = article_name_cell
            article_name = current_article_name
            metric_code = article_sheet.cell(row=row, column=2).value
            metric_label = article_sheet.cell(row=row, column=3).value
            plan = cell_import_value(article_sheet.cell(row=row, column=4))
            fact = cell_import_value(article_sheet.cell(row=row, column=5))
            comment = article_sheet.cell(row=row, column=6).value
            if not article_name and not metric_code and not plan and not fact and not comment:
                continue
            article_entries.append(
                {
                    "article_name": article_name,
                    "metric_code": metric_code,
                    "metric_label": metric_label,
                    "plan": plan,
                    "fact": fact,
                    "comment": comment,
                }
            )

    return {"summary_metrics": metrics, "article_entries": article_entries}
